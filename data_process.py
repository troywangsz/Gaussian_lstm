import torch
from torch import nn
from torch.autograd import Variable
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

wb = openpyxl.load_workbook(r"C:/Users/61083/Desktop/HIT/智能温室/py/rnn/data/rnn_data.xlsx")
now_sheet = wb[wb.sheetnames[0]]
sample = []
label = []

TAGP_max = 0
temp_min = 30
temp_max = -1
batch_num = 160
for year in range(0, 8):
    now_sheet = wb[wb.sheetnames[year]]
    row_index = 2
    batch_now = 0
    LAI_max = 0
    # seq_min = 200
    # seq_max = -1
    while True:
        sample_tmp = []
        label_tmp = []
        while(now_sheet.cell(row=row_index, column=1).value != 'Divided'):
            val_tmp1 = []
            val_tmp2 = []
            LAI_tmp = now_sheet.cell(row=row_index, column=1).value
            LAI_max = max(LAI_max, LAI_tmp)
            val_tmp1.append(LAI_tmp)

            temp_tmp = now_sheet.cell(row=row_index, column=2).value
            temp_min = min(temp_min,temp_tmp)
            val_tmp1.append(temp_tmp)

            temp_tmp = now_sheet.cell(row=row_index, column=3).value
            temp_max = max(temp_max, temp_tmp)
            val_tmp1.append(temp_tmp)

            IRRAD_tmp = now_sheet.cell(row=row_index, column=4).value
            val_tmp1.append(IRRAD_tmp)

            VAP_tmp = now_sheet.cell(row=row_index, column=5).value
            val_tmp1.append(VAP_tmp)

            WIND_tmp = now_sheet.cell(row=row_index, column=6).value
            val_tmp1.append(WIND_tmp)

            RAIN_tmp = now_sheet.cell(row=row_index, column=7).value
            val_tmp1.append(RAIN_tmp)

            LAI_tmp = now_sheet.cell(row=row_index, column=8).value
            LAI_max = max(LAI_max, LAI_tmp)
            val_tmp2.append(LAI_tmp)

            # TAGP_tmp = now_sheet.cell(row=row_index, column=9).value
            # TAGP_max = max(TAGP_max, TAGP_tmp)
            # val_tmp2.append(TAGP_tmp)
            sample_tmp.append(val_tmp1)
            label_tmp.append(val_tmp2)
            row_index += 1
        # seq_max = max(seq_max, len(sample_tmp))
        # seq_min = min(seq_min, len(sample_tmp))
        if(len(sample_tmp) == 199):
            sample.append(sample_tmp)
            label.append(label_tmp)
        if(now_sheet.cell(row=row_index+1, column=1).value == None):
            break
        batch_now += 1
        if(batch_now >= batch_num):
            break
        row_index += 1

sample = np.array(sample)
label = np.array(label)
temp_normal = temp_max-temp_min

for batch in sample:
    for seq in batch:
        seq[0] = seq[0]/LAI_max
        seq[1] = (seq[1]-temp_min)/temp_normal
        seq[2] = (seq[2] - temp_min) / temp_normal
for batch in label:
    for seq in batch:
        seq[0] = seq[0] / LAI_max
        # seq[1] = seq[1] / TAGP_max
        # seq[0] = seq[0]/TAGP_max

# print(sample.shape)
# print(label.shape)

sample = sample.reshape((-1,7))
label = label.reshape((-1,1))

print(sample.shape)
print(label.shape)

data_all = np.concatenate((sample,label),axis=1)
print(data_all.shape)

# data_all = pd.DataFrame(data_all,columns=["LAI_real","Temp_min","Temp_max","Irrad","Vap","Wind","Rain","TAGP_pre"])
data_all = pd.DataFrame(data_all,columns=["LAI_real","Temp_min","Temp_max","Irrad","Vap","Wind","Rain","LAI_pre"])


data_all.to_csv("data/1043_199_7_1_sug_LAI.csv",index=False,na_rep=0.0)






